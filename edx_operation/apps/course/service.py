from collections import defaultdict
from contextlib import suppress
from copy import deepcopy
import hashlib
import json
import math
from pathlib import Path
import random
import re
import string
import tarfile
import tempfile
from urllib.parse import parse_qs, urlparse
from uuid import uuid4

from django.core.exceptions import ValidationError
from lxml import etree
from opaque_keys.edx.locator import LibraryLocator
import openpyxl

import xlrd


class CourseCreator:
    def __init__(
        self,
        org,
        number,
        run,
        course_document,
        cover,
        time_document,
        content_document,
        start,
        end,
        enrollment_start,
        enrollment_end,
        certificate_available_date,
    ):
        self.org = org
        self.number = number
        self.run = run
        self.course_document = course_document
        self.cover = cover
        self.time_document = time_document
        self.content_document = content_document
        self.start = start
        self.end = end
        self.enrollment_start = enrollment_start
        self.enrollment_end = enrollment_end
        self.certificate_available_date = certificate_available_date

    def create_course(self):
        # return course.unique.tar.gz
        return self._create_course()

    def _create_course(self):
        self._load_olx_data()

        # course 생성/업로드
        with tempfile.TemporaryDirectory() as tmpdir:
            # policies
            policies_dir = Path(tmpdir) / "policies" / self.run
            policies_dir.mkdir(parents=True, exist_ok=True)

            # course root olx
            course_root = self.course_olx

            with open(policies_dir / "policy.json", "w", encoding="utf8") as f:
                policy = {
                    f"course/{self.run}": {
                        "tabs": [
                            {
                                "course_staff_only": False,
                                "name": "Home",
                                "type": "course_info",
                            },
                            {
                                "course_staff_only": False,
                                "name": "Course",
                                "type": "courseware",
                            },
                            {
                                "course_staff_only": False,
                                "name": "강의 계획서",
                                "type": "static_tab",
                                "url_slug": "syllubus",
                            },
                            {
                                "course_staff_only": False,
                                "name": "Textbooks",
                                "type": "textbooks",
                            },
                            {
                                "course_staff_only": False,
                                "name": "Discussion",
                                "type": "discussion",
                            },
                            {
                                "course_staff_only": False,
                                "name": "Wiki",
                                "type": "wiki",
                            },
                            {
                                "course_staff_only": False,
                                "name": "Progress",
                                "type": "progress",
                            },
                            {
                                "course_staff_only": False,
                                "name": "Dates",
                                "type": "dates",
                            },
                            {
                                "course_staff_only": False,
                                "name": "Teams",
                                "type": "teams",
                            },
                        ],
                        "teams_configuration": {
                            "topics": [
                                {
                                    "description": "교수자가 지정하는 주제에 대해서 팀을 만들고 가입할 수 있습니다.",
                                    "name": "교수자 지정 토론 주제",
                                    "id": "team_instructional_topic",
                                },
                                {
                                    "description": "학습자가 주제를 정하여 팀을 만들고 가입할 수 있습니다.",
                                    "name": "자율 토론 주제",
                                    "id": "team_free_topic",
                                },
                            ],
                            "max_team_size": 499,
                        },
                    }
                }
                f.write(json.dumps(policy, indent=4))

            progress_count = len(course_root.xpath('//sequential[@format="진도"]'))
            midterm_exam_count = len(course_root.xpath('//sequential[@format="진행평가"]'))
            final_exam_count = len(course_root.xpath('//sequential[@format="시험"]'))
            report_exam_count = len(course_root.xpath('//sequential[@format="과제"]'))

            with open(policies_dir / "grading_policy.json", "w", encoding="utf8") as f:
                grading_policy = {
                    "GRADER": [
                        {
                            "drop_count": 0,
                            "min_count": progress_count,
                            "short_label": "진도",
                            "type": "진도",
                            "weight": 0.0,
                        },
                        {
                            "drop_count": 0,
                            "min_count": midterm_exam_count,
                            "short_label": "진행평가",
                            "type": "진행평가",
                            "weight": 0.1,
                        },
                        {
                            "drop_count": 0,
                            "min_count": final_exam_count,
                            "short_label": "시험",
                            "type": "시험",
                            "weight": 0.6,
                        },
                        {
                            "drop_count": 0,
                            "min_count": report_exam_count,
                            "short_label": "과제",
                            "type": "과제",
                            "weight": 0.3,
                        },
                    ],
                    "GRADE_CUTOFFS": {"Pass": 0.6},
                }
                if not report_exam_count:
                    grading_policy["GRADER"][2]["weight"] = 0.9
                    del grading_policy["GRADER"][3]
                f.write(json.dumps(grading_policy, indent=4))

            # tabs: syllubus
            tabs_dir = Path(tmpdir) / "tabs"
            tabs_dir.mkdir(parents=True, exist_ok=True)
            etree.ElementTree(self.syllubus_olx).write(
                f"{tabs_dir}/syllubus.html", encoding="utf8", pretty_print=True
            )

            # static
            static_dir = Path(tmpdir) / "static"
            static_dir.mkdir(parents=True, exist_ok=True)
            with open(f"{static_dir}/cover.png", "wb") as f:
                # with Pillow
                with self.cover.open("rb") as image_file:
                    f.write(image_file.read())

            # course about
            about_dir = Path(tmpdir) / "about"
            about_dir.mkdir(parents=True, exist_ok=True)
            with open(f"{about_dir}/effort.html", "w") as f:
                effort = len(course_root.xpath('//sequential[@format="진도" or @format="시험"]'))
                f.write(f"{effort}H")
            with open(f"{about_dir}/entrance_exam_enabled.html", "w") as f:
                f.write("false")
            with open(f"{about_dir}/entrance_exam_minimum_score_pct.html", "w") as f:
                f.write("50")

            # html content
            html_dir = Path(tmpdir) / "html"
            html_dir.mkdir(parents=True, exist_ok=True)
            for html_div in course_root.xpath("//html/div"):
                html = html_div.getparent()
                filename = f"{html.get('filename')}.html"
                etree.ElementTree(html_div).write(
                    f"{html_dir}/{filename}", encoding="utf8", pretty_print=True
                )
                html.remove(html_div)

            # component > vertical > sequential > chapter
            CourseCreator._write_elements(course_root.xpath("//html"), tmpdir)
            CourseCreator._write_elements(course_root.xpath("//problem"), tmpdir)
            CourseCreator._write_elements(course_root.xpath("//library_content"), tmpdir)
            CourseCreator._write_elements(course_root.xpath("//vertical"), tmpdir)
            CourseCreator._write_elements(course_root.xpath("//sequential"), tmpdir)
            CourseCreator._write_elements(course_root.xpath("//chapter"), tmpdir)
            CourseCreator._write_elements(
                course_root.xpath("//course/course"), tmpdir, create_link=False
            )

            # course root
            etree.ElementTree(course_root).write(
                f"{tmpdir}/course.xml", encoding="utf8", pretty_print=True
            )

            # Create the tar.gz file
            file = f"{tmpdir}/course.{''.join(random.choices(string.ascii_letters, k=8))}.tar.gz"
            with tarfile.open(file, "w:gz") as tar:
                # Add all files from the tmpdir_path to the tarball
                tar.add(tmpdir, arcname="course")

            return open(file, "rb")

    def _load_olx_data(self):
        wb = CourseCreator._load_workbook(self.course_document)
        self.course_workbook = wb

        self.course_sheet = wb.worksheets[0]
        self.course_title = self.course_sheet["B2"].value

        # tranform workbook structure
        self._uniform_course_workbook()

        # libraries
        self.mid_exam_library_olx = self._create_library_olx("mid_exam", wb.worksheets[2])
        self.final_exam_library_olx = self._create_library_olx("final_exam", wb.worksheets[3])
        try:
            # 과제는 없을 수도 있음
            self.report_exam_library_olx = self._create_library_olx(
                "report_exam", wb.worksheets[4]
            )
        except IndexError:
            self.report_exam_library_olx = None

        # course
        self.course_olx = self._create_course_olx()

        # syllubus
        self.syllubus_olx = self._create_syllubus()

    def _uniform_course_workbook(self):
        # sheet2 mid_exam, sheet3 final_exam
        for sheet in [2, 3]:
            try:
                exam_sheet = self.course_workbook.worksheets[sheet]
            except IndexError:
                break

            if not exam_sheet["I3"].value:
                raise ValueError(f"{exam_sheet.title}(I3) 문서 형식이 잘못되었습니다.")

            # transform 4 multiple selection to 5 multiple selection
            if exam_sheet["I3"].value.replace(" ", "") == "정답":
                exam_sheet.insert_cols(9)

        # report sheet
        try:
            report_sheet = self.course_workbook.worksheets[4]
            # data starts from row 4
            report_sheet.insert_rows(2)
            # problem points
            report_sheet.insert_cols(5, 5)
            # explaination
            report_sheet.insert_cols(12)
        except IndexError:
            pass

    def _create_course_olx(self):
        def build_course_tree(course):
            outline_numberings = [c.value for c in self.course_sheet["B"] if c.value]
            outline_len = float(str(outline_numberings[-1]).replace("차시", "").strip())

            # completion seconds
            completion_seconds = self._completion_seconds()
            if completion_seconds and outline_len != len(completion_seconds):
                raise ValueError("과정개요서와 학습시간 계획서가 일치하지 않습니다.")

            # content url
            content_urls = self._content_urls()
            if content_urls and outline_len != len(content_urls):
                raise ValueError("과정개요서와 콘텐츠 경로 문서가 일치하지 않습니다.")

            # course tree
            for i, row in enumerate(self.course_sheet.iter_rows(min_row=5)):
                # 4 sequentials in one chapter
                sequentials_in_chapter = 4
                chapter_index, remain = divmod(i, sequentials_in_chapter)
                chapter_display_name = f"섹션 {chapter_index + 1}"

                # sequential title
                sequential_title = row[2].value

                # empty line
                if sequential_title is None:
                    break

                # next chapter
                if remain == 0:
                    etree.SubElement(
                        course,
                        "chapter",
                        url_name=uuid4().hex,
                        display_name=chapter_display_name,
                        highlights="[]",
                    )

                chapter = course.xpath("//chapter")[-1]
                chapter.set(
                    "highlights",
                    json.dumps(json.loads(chapter.get("highlights")) + [sequential_title]),
                )

                # sequential
                sequential = etree.SubElement(
                    chapter,
                    "sequential",
                    url_name=uuid4().hex,
                    display_name=sequential_title,
                    exam_review_rules="",
                    format="진도",
                    graded="true",
                    show_correctness="past_due",
                    is_onboarding_exam="false",
                    is_practice_exam="false",
                    is_proctored_enabled="false",
                    is_time_limited="false",
                    default_time_limit_minutes="0",
                )

                # vertical
                vertical = etree.SubElement(
                    sequential,
                    "vertical",
                    url_name=uuid4().hex,
                    display_name="학습 하기",
                )

                # html
                html_url_name = uuid4().hex
                html_component = etree.SubElement(
                    vertical,
                    "html",
                    url_name=html_url_name,
                    filename=html_url_name,
                    display_name="학습 안내",
                    editor="raw",
                )

                # completion info
                sequential_completion_seconds = completion_seconds.get(i + 1) or 1560
                completion_delay_seconds = math.ceil(sequential_completion_seconds / 2)
                completion_info_html = f"""
                    <div>
                        <h3>학습 안내</h3>
                        <ol>
                            <li>
                                이번 섹션의 예상 학습 시간은
                                <strong>{math.ceil(sequential_completion_seconds / 60)}분</strong>이고,
                                최소 학습 시간은 <strong>{math.ceil(completion_delay_seconds / 60)}분</strong>입니다.
                            </li>
                            <li>아래 내용을 최소 학습 시간 이상으로 학습하십시오.</li>
                        </ol>
                    </div>
                """
                html_component.append(etree.XML(completion_info_html))

                # embed content with completion info
                content_url = content_urls.get(i + 1)
                if not content_url:
                    try:
                        content_url = str(row[15].value)
                    except IndexError:
                        content_url = None

                    # sample
                    if not content_url:
                        content_url = "https://youtu.be/NHwxTZmdxKw"

                youtube_id = CourseCreator._extract_youtube_id(content_url)

                if youtube_id or re.search(r"\.mp4$|\.m3u8$|\.ogg$|\.webm$|\.mpd$", content_url):
                    video_component = etree.SubElement(
                        vertical,
                        "video",
                        url_name=uuid4().hex,
                        display_name="동영상 학습",
                        completion_delay_seconds=str(completion_delay_seconds),
                        edx_video_id="",
                        html5_sources=json.dumps([content_url]),
                        youtube_id_1_0=youtube_id or "",
                    )
                    etree.SubElement(video_component, "source", src=content_url)
                else:
                    html_url_name = uuid4().hex
                    html_component = etree.SubElement(
                        vertical,
                        "html",
                        url_name=html_url_name,
                        filename=html_url_name,
                        display_name="학습 하기",
                        completion_delay_seconds=str(completion_delay_seconds),
                        editor="raw",
                    )
                    html_div = etree.SubElement(
                        etree.SubElement(html_component, "div"),
                        "div",
                    )

                    # ex. youtube iframe
                    if content_url.lstrip().startswith("<"):
                        parsed_html = etree.fromstring(content_url, parser=etree.HTMLParser())
                        html_div.append(parsed_html)
                    else:
                        etree.SubElement(
                            html_div,
                            "iframe",
                            src=content_url,
                            allow="fullscreen *; encrypted-media *",
                            scrolling="no",
                            style="border: none; overflow: hidden;",
                            frameborder="0",
                            allowfullscreen="true",
                            webkitallowfullscreen="true",
                            mozAllowFullScreen="true",
                        )

                # discussion
                etree.SubElement(
                    vertical,
                    "discussion",
                    url_name=uuid4().hex,
                    **{"xblock-family": "xblock.v1"},
                    display_name="질문 답변 / 학습 토론",
                    discussion_category=chapter_display_name,
                    discussion_target=sequential_title,
                )

        def insert_wiki(course, slug):
            # course wiki
            etree.SubElement(course, "wiki", slug=slug)

        def insert_survey_chapter(course):
            # 설문 조사/수강 후기 추가
            final_exam_chapter = course.xpath("//chapter[@display_name='시험']")
            if final_exam_chapter:
                # sequential
                sequential = etree.Element(
                    "sequential",
                    url_name=uuid4().hex,
                    display_name="설문 조사 / 수강 후기",
                )
                final_exam_chapter[-1].insert(0, sequential)

                # vertical
                vertical = etree.SubElement(
                    sequential,
                    "vertical",
                    url_name=uuid4().hex,
                    display_name="설문 조사 / 수강 후기",
                )

                # 설문 조사
                survey_questions = [
                    [uuid4().hex[:10], {"img": "", "img_alt": "", "label": question}]
                    for question in (
                        "수강한 과정은 최신 동향을 잘 반영하였다.",
                        "수강한 과정은 실제 업무에 도움이 되었다.",
                        "제출한 과제에 대해 적절한 평가와 피드백이 이루어졌다.",
                        "교육내용은 나의 업무능력개발에 도움이 될 것이다.",
                        "학습자의 질문에 신속하고 적절하게 답변했다.",
                        "학습 중 불편사항(프로그램 오류 등)발생 시 적절한 해결방안을 제시했다.",
                        "과정안내, 수료기준, 프로그램 안내를 적절하게 제공했다.",
                        "홈페이지 및 시스템은 학습에 필요한 편의기능을 제공했다.",
                        "운영자의 학습독려(sns, 이메일 등)는 학습진행에 도움을 주었다.",
                        "교육내용은 명확하고 이해하기 쉽게 표현되었다.",
                        "학습분량은 적절했다.",
                        "수강한 과정에 만족한다.",
                    )
                ]
                survey_answers = [
                    [uuid4().hex[:10], answer]
                    for answer in (
                        "절대아니다",
                        "아니다",
                        "보통이다",
                        "그렇다",
                        "매우그렇다",
                    )
                ]
                survey_feedback = "#### 설문 조사에 응답해 주셔서 감사합니다.\n더 나은 과정을 제공할 수 있도록 최선을 다 하겠습니다."

                # component survey
                etree.SubElement(
                    vertical,
                    "survey",
                    url_name=uuid4().hex,
                    **{"xblock-family": "xblock.v1"},
                    feedback=survey_feedback,
                    max_submissions="1",
                    private_results="false",
                    answers=json.dumps(survey_answers),
                    block_name="과정 만족도 설문 조사",
                    questions=json.dumps(survey_questions),
                )

                # 수강 후기
                etree.SubElement(
                    vertical,
                    "discussion",
                    url_name=uuid4().hex,
                    **{"xblock-family": "xblock.v1"},
                    display_name="수강 후기",
                    discussion_category="수강 후기",
                    discussion_target="수강 후기",
                )

        def insert_library_chapter(course, libraries):
            for library, title in libraries:
                if library is None:
                    continue

                # chapter
                chapter = etree.Element(
                    "chapter",
                    url_name=uuid4().hex,
                    display_name=title,
                )

                # 진행평가는 중간에 삽입
                if title == "진행평가":
                    course.insert(math.ceil(len(course.xpath("//chapter")) / 2.0) + 1, chapter)
                else:
                    course.append(chapter)

                # sequential
                # 과정개요서 상으로 진행평가/과제/시험은 항상 1개
                sequential = etree.SubElement(
                    chapter,
                    "sequential",
                    url_name=uuid4().hex,
                    display_name=title,
                    exam_review_rules="",
                    format=title,
                    graded="true",
                    show_correctness="past_due",
                    is_onboarding_exam="false",
                    is_practice_exam="false",
                    is_proctored_enabled="false",
                    is_time_limited="true" if title == "시험" else "false",
                    default_time_limit_minutes="60",
                )

                # vertical
                vertical = etree.SubElement(
                    sequential,
                    "vertical",
                    url_name=uuid4().hex,
                    display_name=f"{title} 응시",
                )

                if title == "진행평가":
                    # 진행평가 유의 사항
                    display_name = "진행평가 응시 유의 사항 안내"
                    info = """
                        <div>
                            <h3>진행평가 응시 유의 사항 안내</h3>
                            <ul>
                                <li>문제를 풀고 문제 별로 [제출] 버튼을 눌러 제출하십시오.</li>
                                <li>모든 문제를 제출 한 후에 학습을 계속 진행하십시오.</li>
                            </ul>
                        </div>
                    """
                elif title == "시험":
                    # 시험 응시 유의 사항
                    display_name = "시험 응시 유의 사항 안내"
                    info = """
                        <div>
                            <h3>시험 응시 유의 사항 안내</h3>
                            <ul>
                                <li>시험 시간은 60분입니다. 답안 제출 후 수정은 불가능 합니다.</li>
                                <li>문제를 풀고 문제 별로 [제출] 버튼을 눌러 제출하십시오. 문제를 모두 제출한 후에는 상단의 [시험 종료] 버튼을 눌러 시험을 종료하십시오.</li>
                                <li>진행 중 창이 닫히거나 컴퓨터가 꺼지는 경우에도 시간 내에는 다시 접속 해서 계속할 수 있습니다.</li>
                                <li>모든 문제를 풀지 못한 채 시간이 종료된 경우에는 답안이 입력된 문제만 자동 제출됩니다.</li>
                            </ul>
                        </div>
                    """
                else:
                    # 과제 응시 유의 사항
                    display_name = "과제 응시 유의 사항 안내"
                    info = """
                        <div>
                            <h3>베낀답안 검사 및 처리 안내</h3>
                            <ul>
                                <li>서술형 문제는 [시스템 베낀답안 검사]와 [교강사 베낀답안 검사]를 실시합니다.</li>
                                <li>[시스템 베낀답안 검사]는 다른 학습자의 답안과 문장, 구문, 단어, 오타의 사용률을 검사하여 80% 이상 일치하면 베낀답안으로 필터링됩니다.</li>
                                <li>[교강사 베낀답안 검사]는 [시스템 베낀답안 검사] 내용과 문제의 성격을 고려하여 최종적으로 베낀답안 여부를 확정합니다.</li>
                                <li>베낀답안으로 확정된 경우에는 미수료로 처리됩니다. 유의하여 주시기 바랍니다.</li>
                            </ul>
                        </div>
                    """

                html_url_name = uuid4().hex
                exam_info = etree.SubElement(
                    vertical,
                    "html",
                    url_name=html_url_name,
                    filename=html_url_name,
                    display_name=display_name,
                    editor="raw",
                )
                exam_info.append(etree.XML(info))

                # 라이브러리 문제 복사
                problem_types = [
                    ("multiplechoiceresponse", "객관식 문제"),
                    ("stringresponse", "단답형 문제"),
                    ("openassessment", "서술형 문제"),
                ]

                # problem typs
                ora_problems = library.xpath("//openassessment")
                string_problems = library.xpath("//stringresponse")
                choice_problems = library.xpath("//multiplechoiceresponse")

                # aggregate
                library_id = LibraryLocator(org=self.org, library=library.attrib["library"])
                kwargs = {}

                if ora_problems:
                    kwargs["ora_problem_count"] = len(ora_problems)
                    kwargs["ora_problem_point"] = int(ora_problems[0].attrib["weight"])
                if string_problems:
                    kwargs["string_problem_count"] = len(string_problems)
                    kwargs["string_problem_point"] = int(
                        string_problems[0].getparent().attrib["weight"]
                    )
                if choice_problems:
                    kwargs["choice_problem_count"] = len(choice_problems)
                    kwargs["choice_problem_point"] = int(
                        choice_problems[0].getparent().attrib["weight"]
                    )
                selection = CourseCreator._calulate_problem_selection(
                    library_id=library_id, **kwargs
                )

                for capa_type, title in problem_types:
                    problem_contents = library.xpath(f"//{capa_type}")
                    if not problem_contents:
                        continue

                    library_content = etree.SubElement(
                        vertical,
                        "library_content",
                        url_name=uuid4().hex,
                        capa_type=capa_type,
                        display_name=title,
                        max_count=str(selection.get(capa_type)),
                        source_library_id=str(library_id),
                    )

                    def generate_id(problem):
                        key_base = "{}:{}:{}".format(
                            str(library_id.for_version(None)).encode("utf-8"),
                            problem.get("url_name"),
                            library_content.get("url_name"),
                        )
                        return hashlib.sha1(key_base.encode("utf-8")).hexdigest()[:20]

                    for problem_content in problem_contents:
                        problem = deepcopy(
                            problem_content
                            # open assessment problem 구조가 다름
                            if capa_type == "openassessment"
                            else problem_content.getparent()
                        )
                        problem.set("url_name", generate_id(problem))
                        library_content.append(problem)

        # create course

        course = etree.Element(
            "course",
            url_name=self.run,
            advanced_modules=json.dumps(["library_content", "survey", "poll"]),
            banner_image="cover.png",
            cert_html_view_enabled="true",
            cert_name_long="수료증",
            cert_name_short="수료증",
            certificate_available_date=self.certificate_available_date,
            certificates=json.dumps(
                {
                    "certificates": [
                        {
                            "course_title": self.course_title,
                            "description": "과정 수료증",
                            "editing": False,
                            "id": random.randint(0, (2**31) - 1),
                            "is_active": True,
                            "name": "수료증",
                            "signatories": [],
                            "version": 1,
                        }
                    ]
                }
            ),
            course_image="cover.png",
            course_visibility="public_outline",
            days_early_for_beta="365.0",
            discussion_sort_alpha="true",
            discussion_topics=json.dumps(
                {
                    "과정 게시판": {
                        "default": True,
                        "id": f"i4x-{self.org}-{self.number}-course-{self.run}",
                    }
                }
            ),
            display_name=self.course_title,
            display_organization=self.org,
            due=self.end,
            enable_subsection_gating="true",
            end=self.end,
            enrollment_start=self.enrollment_start,
            enrollment_end=self.enrollment_end,
            graceperiod="",
            highlights_enabled_for_messaging="false",
            instructor_info=json.dumps({"instructors": []}),
            invitation_only="true",
            language="ko",
            max_attempts="1",
            mobile_available="true",
            other_course_settings=json.dumps({}),
            showanswer="past_due",
            start=self.start,
            license="all-rights-reserved",
            user_partitions="[]",
        )

        # course
        build_course_tree(course)

        # course wiki
        slug = f"{self.org}.{self.number}.{self.run}"
        insert_wiki(course, slug)

        # insert library
        libraries = [
            (self.mid_exam_library_olx, "진행평가"),
            (self.final_exam_library_olx, "시험"),
        ]
        if self.report_exam_library_olx is not None:
            libraries.append((self.report_exam_library_olx, "과제"))

        insert_library_chapter(course, libraries)

        # survey
        insert_survey_chapter(course)

        # course root
        course_root = etree.Element("course", url_name=self.run, org=self.org, course=self.number)
        course_root.append(course)

        return course_root

    def _create_syllubus(self):
        syllubus = etree.Element("div")
        etree.SubElement(syllubus, "h2").text = self.course_title
        syllubus_sequentials = etree.SubElement(syllubus, "ol")

        # completion seconds
        completion_seconds = self._completion_seconds()

        for i, row in enumerate(self.course_sheet.iter_rows(min_row=5)):
            # syllubus
            syllubus_sequential = etree.SubElement(syllubus_sequentials, "li")
            syllubus_sequential_title = etree.SubElement(syllubus_sequential, "h3")

            # sequential title
            sequential_title = row[2].value
            ncs_code = row[7].value
            ncs_sub_code = row[8].value
            objectives = str(row[3].value)
            training_content = str(row[4].value)
            main_activities = str(row[6].value)

            # 학습 시간 표시
            syllubus_sequential_title.text = sequential_title
            span = etree.SubElement(syllubus_sequential_title, "span")
            seconds = completion_seconds.get(i + 1) or 1560
            span.text = f"{math.ceil(seconds / 60)}분" if seconds else ""

            # ncs unit 표시
            if ncs_code and ncs_code.lower().replace(" ", "") != "비ncs":
                ncs_code = ncs_code.replace("\n", " ")
                etree.SubElement(syllubus_sequential, "span").text = f"NCS 능력단위 {ncs_code}"

            if ncs_sub_code and ncs_sub_code.lower().replace(" ", "") != "비ncs":
                ncs_sub_code = ncs_sub_code.replace("\n", " ")
                etree.SubElement(syllubus_sequential, "span").text = f"NCS 능력단위요소 {ncs_sub_code}"

            # 학습 목표
            sequential_objects = etree.SubElement(syllubus_sequential, "ol")
            etree.SubElement(sequential_objects, "h4").text = "학습 목표"
            for objective in objectives.split("\n"):
                objective = re.sub(r"^\d+\.?\s", "", objective.strip())
                if not objective:
                    continue
                etree.SubElement(sequential_objects, "li").text = objective

            # 학습 내용
            sequential_contents = etree.SubElement(syllubus_sequential, "ol")
            etree.SubElement(sequential_contents, "h4").text = "학습 내용"
            for cotent in training_content.split("\n"):
                cotent = re.sub(r"^\d+\.?\s", "", cotent.strip())
                if not cotent:
                    continue
                etree.SubElement(sequential_contents, "li").text = cotent

            # 주요 학습활동
            sequential_activity = etree.SubElement(syllubus_sequential, "ol")
            etree.SubElement(sequential_activity, "h4").text = "주요 학습활동"

            for activity in main_activities.split("\n"):
                activity = re.sub(r"^\d+\.?\s", "", activity.strip())
                if not activity:
                    continue
                etree.SubElement(sequential_activity, "li").text = activity

        return syllubus

    def _completion_seconds(self):
        sequentials = defaultdict(int)
        if not self.time_document:
            return sequentials

        wb = CourseCreator._load_workbook(self.time_document)
        ws = wb.worksheets[0]

        sequential_index_cache = 0

        for row in ws.iter_rows(min_row=7):
            if not row[3].value and not row[4].value:
                break
            if re.search(r"합계$|총계$", str(row[3].value)):
                continue

            sequential_index = row[1].value
            if sequential_index:
                sequential_index_cache = int(sequential_index)
            completion_minutes = int(row[7].value or 0)
            completion_seconds = int(row[8].value or 0)
            sequentials[sequential_index_cache] += completion_minutes * 60 + completion_seconds
        return sequentials

    def _content_urls(self):
        content_urls = {}
        if not self.content_document:
            return content_urls

        wb = CourseCreator._load_workbook(self.content_document)
        ws = wb.worksheets[0]

        for row in ws.iter_rows(min_row=2):
            url = row[2].value
            if not url:
                break

            sequential_index = int(row[1].value)
            content_urls[sequential_index] = url

        return content_urls

    def _create_library_olx(self, library_type, library_sheet):
        library_code = f"{self.number}-{library_type}"

        # library root
        library = etree.Element(
            "library",
            url_name="library",
            library=library_code,
            **{"xblock-family": "xblock.v1"},
            display_name=f"{self.course_title} {self.org}+{library_code}",
            org=self.org,
        )

        # build problem
        for row in library_sheet.iter_rows(min_row=4):
            # 문제 정보
            question = row[3].value
            if not question:
                # 건너뛰지 않고 break
                break

            # sequential
            sequential = row[0].value
            if isinstance(sequential, float):
                sequential = int(sequential)

            # ncs_code, ncs_sub_code
            try:
                ncs_code = row[14].value
            except IndexError:
                ncs_code = ""
            try:
                ncs_sub_code = str(row[15].value or "")
            except IndexError:
                ncs_sub_code = ""

            # correct_answers
            correct_answers = row[9].value or ""
            if isinstance(correct_answers, float):
                correct_answers = int(correct_answers)

            # 정답을 O/X로 표시한 경우
            if isinstance(correct_answers, str):
                correct_answers = correct_answers.strip()
                if correct_answers.lower() == "o":
                    correct_answers = 1
                elif correct_answers.lower() == "x":
                    correct_answers = 2

            correct_answers = [x.strip() for x in re.split(r",|\n", str(correct_answers)) if x]

            # explanation
            explanation = row[10].value or ""

            # fix 100점 -> 100
            try:
                possible_score = int(re.search(r"\d+", str(row[12].value))[0])
            except TypeError:
                # 4지선다 L, 5지선다 M
                raise ValueError(
                    f"문제 배점 오류 {library_sheet.title} {row[11].coordinate} 또는 {row[12].coordinate}"
                )

            problem_type = row[1].value
            if not problem_type:
                raise ValueError(f"문제 유형 오류 {library_sheet.title} {row[1].coordinate}")

            # ora problem
            if re.search(r"서술형|적응형|실습형", problem_type):
                # 서술형 평가 정답 및 해설을 한칸씩 밀린 경우
                # 서술형 평가는 복수 정답이 없음
                correct_answers = row[9].value
                if not correct_answers and explanation:
                    correct_answers = explanation
                if not explanation or correct_answers == explanation:
                    explanation = row[11].value or ""

                problem = CourseCreator._create_ora_problem(
                    question,
                    correct_answers,
                    explanation,
                    possible_score,
                    sequential,
                    ncs_code,
                    ncs_sub_code,
                )

            # capa problem
            else:
                # choices에서 뒤로부터 None 제거
                choices = [row[i].value for i in range(4, 9)]
                while choices and not choices[-1]:
                    del choices[-1]

                problem = CourseCreator._create_capa_problem(
                    question,
                    correct_answers,
                    explanation,
                    possible_score,
                    choices,
                    problem_type,
                    sequential,
                    ncs_code,
                    ncs_sub_code,
                )

            library.append(problem)

        # 문제가 하나도 없을 때
        if not library.xpath("./*"):
            return None

        return library

    @staticmethod
    def _extract_youtube_id(url, ignore_playlist=False):
        urls = urlparse(url)
        if urls.hostname == "youtu.be":
            return urls.path[1:]
        if urls.hostname in {"www.youtube.com", "youtube.com", "music.youtube.com"}:
            if not ignore_playlist:
                # use case: get playlist id not current video in playlist
                with suppress(KeyError):
                    return parse_qs(urls.query)["list"][0]
            if urls.path == "/watch":
                return parse_qs(urls.query)["v"][0]
            if urls.path[:7] == "/watch/":
                return urls.path.split("/")[1]
            if urls.path[:7] == "/embed/":
                return urls.path.split("/")[2]
            if urls.path[:3] == "/v/":
                return urls.path.split("/")[2]
        return None

    @staticmethod
    def _create_ncs_tip(sequential, ncs_code, ncs_sub_code):
        tip = etree.Element("span")

        if sequential:
            location = int(sequential) if isinstance(sequential, float) else sequential
            etree.SubElement(tip, "span").text = f"학습 위치: {location} 차시"

        newline = "\n"
        if ncs_code and ncs_code.lower().replace(" ", "") != "비ncs":
            etree.SubElement(
                tip,
                "span",
            ).text = f'NCS 능력단위: {ncs_code.replace(newline, " ")}'

        if ncs_sub_code and ncs_sub_code.lower().replace(" ", "") != "비ncs":
            etree.SubElement(
                tip,
                "span",
            ).text = f'NCS 능력단위 요소: {ncs_sub_code.replace(newline, " ")}'

        return tip

    @staticmethod
    def _create_ora_problem(
        question,
        correct_answers,
        explanation,
        possible_score,
        sequential,
        ncs_code,
        ncs_sub_code,
    ):
        ora_template = """
            <openassessment url_name="6a7c1d51be1449a38fa6f430c4f63109"
                    display_name="서술형 문제"
                    submission_start="2023-01-01T00:00:00+09:00"
                    submission_due="2029-12-31T23:59:59+09:00"
                    text_response="required"
                    text_response_editor="text"
                    file_upload_response="optional"
                    file_upload_type="custom"
                    white_listed_file_types="txt,hwp,doc,docx,rtf,xls,xlsx,pdf.png,jpg"
                    allow_multiple_files="True"
                    allow_latex="False"
                    prompts_type="html"
                    teams_enabled="False"
                    selected_teamset_id=""
                    enable_staff_override_point="True"
                    show_rubric_during_response="False">
                <title>서술형 문제</title>
                <assessments>
                    <assessment name="staff-assessment" enable_flexible_grading="False" required="True"/>
                </assessments>
                <prompts>
                    <prompt>
                        <description>여기에 문제</description>
                    </prompt>
                </prompts>
                <rubric>
                    <criterion>
                        <name>0</name>
                        <label>평가 기준</label>
                        <prompt>평가 기준 상세</prompt>
                        <option points="0">
                            <name>0</name>
                            <label>베낀답안</label>
                            <explanation>베낀답안에 해당되어 0점 처리</explanation>
                        </option>
                        <option points="0">
                            <name>1</name>
                            <label>아주 나쁨</label>
                            <explanation>주어진 문제와 전혀 상관 없는 답변</explanation>
                        </option>
                        <option points="1">
                            <name>2</name>
                            <label>나쁨</label>
                            <explanation>주어진 문제와 상관 없는 답변</explanation>
                        </option>
                        <option points="2">
                            <name>3</name>
                            <label>보통</label>
                            <explanation>보통 수준의 답변</explanation>
                        </option>
                        <option points="3">
                            <name>4</name>
                            <label>좋음</label>
                            <explanation>문제에 대한 적절한 답변</explanation>
                        </option>
                        <option points="4">
                            <name>5</name>
                            <label>아주 좋음</label>
                            <explanation>문제에 대한 정확한 답변</explanation>
                        </option>
                        <option points="5">
                            <name>staff_override_point</name>
                            <label>교강사 평가 기준 적용</label>
                            <explanation>-</explanation>
                        </option>
                    </criterion>
                    <feedbackprompt>채점자 피드백</feedbackprompt>
                    <feedback_default_text>제출한 답안을 평가 기준 세부 항목 별로 채점했습니다.</feedback_default_text>
                </rubric>
            </openassessment>
        """
        ora = etree.XML(ora_template)
        ora.set("url_name", uuid4().hex)
        ora.set("weight", str(possible_score))

        # trick starting white space: fix encoding error
        question = f""" <pre><span>{question}</span></pre>"""
        tip = CourseCreator._create_ncs_tip(sequential, ncs_code, ncs_sub_code)
        ora.xpath("//description")[0].text = question + etree.tostring(tip, encoding="unicode")

        options = ora.xpath("//criterion/option")
        partition_count = len(options) - 1
        score_partition = [
            round(i * float(possible_score) / (partition_count - 1))
            for i in range(0, partition_count)
        ]
        for i, score_part in enumerate(score_partition, start=1):
            options[i].set("points", str(score_part))

        for explanation_tag in ora.xpath("//option/explanation")[1:]:
            # trick starting white space: fix encoding error
            explanation_tag.text = f" {explanation}"
        ora.xpath("//feedback_default_text")[0].text = correct_answers

        return ora

    @staticmethod
    def _create_capa_problem(
        question,
        correct_answers,
        explanation,
        possible_score,
        choices,
        problem_type,
        sequential,
        ncs_code,
        ncs_sub_code,
    ):
        problem = etree.Element(
            "problem",
            url_name=uuid4().hex,
            display_name=f'{problem_type.replace(" 문제", "")} 문제',
            markdown="null",
            max_attempts="1",
            show_reset_button="false",
            showanswer="past_due",
            submission_wait_seconds="0",
            weight=str(possible_score),
        )

        if choices:
            problem_content = etree.SubElement(problem, "multiplechoiceresponse")
        else:
            problem_content = etree.SubElement(
                problem, "stringresponse", answer=correct_answers[0], type="ci"
            )

        # question
        label = etree.SubElement(problem_content, "label")
        etree.SubElement(label, "h3").text = str(question)

        # tip
        tip = CourseCreator._create_ncs_tip(sequential, ncs_code, ncs_sub_code)
        etree.SubElement(problem_content, "description").append(tip)

        # choices / input box
        if choices:
            choicegroup = etree.SubElement(problem_content, "choicegroup", type="MultipleChoice")
            for i, choice in enumerate(choices, start=1):
                etree.SubElement(
                    choicegroup,
                    "choice",
                    correct="true" if str(i) in correct_answers else "false",
                ).text = str(choice)
        else:
            for correct_answer in correct_answers[1:]:
                etree.SubElement(problem_content, "additional_answer", answer=correct_answer)
            etree.SubElement(problem_content, "textline", size="60")

        # explanation
        solution = etree.SubElement(problem_content, "solution")
        solution_div = etree.SubElement(solution, "div")
        etree.SubElement(solution_div, "p").text = "설명"
        etree.SubElement(solution_div, "p").text = explanation

        return problem

    @staticmethod
    def _calulate_problem_selection(
        ora_problem_count=0,
        ora_problem_point=1,
        string_problem_count=0,
        string_problem_point=1,
        choice_problem_count=0,
        choice_problem_point=1,
        library_id=None,
    ):
        # 문제 은행에서 여러 문제 유형을 섞어서 100점에 맞게 문제 추출

        selection_limit = 1 / 3
        if library_id and "mid_exam" in library_id.library:
            selection_limit = 1 / 1

        # 과제, 시험에서 문제유형이 3개 이하인 경우
        # 1 / 3 limit을 적용하지 하지 않고 1 / 1을 적용
        def caculate_limit(count):
            limit = count * selection_limit
            return 1 if 0 < limit < 1 else int(limit)

        o_problem_limit = caculate_limit(ora_problem_count)
        o_problem_max_point = o_problem_limit * ora_problem_point
        s_problem_limit = caculate_limit(string_problem_count)
        s_problem_max_point = s_problem_limit * string_problem_point
        c_problem_limit = caculate_limit(choice_problem_count)
        c_problem_max_point = c_problem_limit * choice_problem_point

        # 문제 갯수
        selections = [
            {
                "openassessment": int(o_part / ora_problem_point),
                "stringresponse": int(s_part / string_problem_point),
                "multiplechoiceresponse": int(c_part / choice_problem_point),
            }
            for o_part in range(o_problem_max_point, -1, -ora_problem_point)
            for s_part in range(s_problem_max_point, -1, -string_problem_point)
            for c_part in range(c_problem_max_point, -1, -choice_problem_point)
            if o_part + s_part + c_part == 100
        ]

        if selections:
            return selections[0]
        else:
            return {
                "openassessment": o_problem_limit,
                "stringresponse": s_problem_limit,
                "multiplechoiceresponse": c_problem_limit,
            }

    @staticmethod
    def _write_elements(elements, root_dir, create_link=True):
        for element in elements:
            tag_dir = Path(root_dir) / element.tag
            tag_dir.mkdir(parents=True, exist_ok=True)
            # save
            url_name = element.attrib.pop("url_name")
            etree.ElementTree(element).write(
                f"{tag_dir}/{url_name}.xml", encoding="utf8", pretty_print=True
            )
            # create link
            parent = element.getparent()
            if create_link:
                link = etree.Element(element.tag, url_name=url_name)
                parent.replace(element, link)
            else:
                parent.remove(element)

    @staticmethod
    def _load_workbook(file):
        # xlsx
        if file.name.endswith(".xlsx"):
            return openpyxl.load_workbook(file, data_only=True)

        if not file.name.endswith(".xls"):
            raise ValidationError("Unknown File Type")

        # xls
        file.seek(0)
        xlsBook = xlrd.open_workbook(file_contents=file.read())
        workbook = openpyxl.Workbook()

        for i in range(0, xlsBook.nsheets):
            xlsSheet = xlsBook.sheet_by_index(i)

            sheet = workbook.active if i == 0 else workbook.create_sheet()
            sheet.title = xlsSheet.name

            for row in range(0, xlsSheet.nrows):
                for col in range(0, xlsSheet.ncols):
                    sheet.cell(row=row + 1, column=col + 1).value = xlsSheet.cell_value(row, col)
        return workbook
